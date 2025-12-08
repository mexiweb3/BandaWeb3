import openpyxl

def debug_columns():
    input_file = "shared/inputs/Spoken - Spaces Dashboard.xlsx"
    print(f"Loading {input_file}...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Check header row (Row 3 assumed)
    print("Header Row (3):")
    for col in range(1, 10):
        cell = ws.cell(row=3, column=col)
        print(f"Col {col}: {cell.value}")

    print("\nChecking first few data rows...")
    for r in range(4, 10):
        c_link = ws.cell(row=r, column=5) # E
        c_prev = ws.cell(row=r, column=4) # D
        
        link_target = c_link.hyperlink.target if c_link.hyperlink else "No Link"
        print(f"Row {r}: Col D='{c_prev.value}', LinkTarget='{link_target}'")

if __name__ == "__main__":
    debug_columns()
