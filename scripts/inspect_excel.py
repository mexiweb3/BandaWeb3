
import openpyxl

file_path = 'shared/inputs/Cohosted - Spaces Dashboard.xlsx'

print(f"Loading {file_path}...")

# Load using openpyxl to access hyperlinks
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Columns: {headers}")

# Inspect first 5 rows
print("\nFirst 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6)):
    row_data = []
    links = []
    for cell in row:
        val = cell.value
        row_data.append(str(val)[:20] + "..." if val and len(str(val))>20 else val)
        if cell.hyperlink:
            links.append(f"Link in col {cell.column}: {cell.hyperlink.target}")
    
    print(f"Row {i+1}: {row_data}")
    if links:
        print(f"  Hyperlinks: {links}")
